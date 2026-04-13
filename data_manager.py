"""SQLiteによるデータ管理とExcel出力"""

from __future__ import annotations

import sqlite3
import io
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DB_PATH = Path(__file__).parent / "companies.db"

COLUMNS = [
    ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
    ("company_name", "TEXT NOT NULL"),
    ("phone", "TEXT"),
    ("address", "TEXT"),
    ("url", "TEXT"),
    ("domain", "TEXT"),
    ("snippet", "TEXT"),
    ("industry", "TEXT"),
    ("product_genre", "TEXT"),
    ("prefecture", "TEXT"),
    ("employee_count", "TEXT"),
    ("employee_range", "TEXT"),
    ("revenue", "TEXT"),
    ("has_overseas", "TEXT"),
    ("notes", "TEXT"),
    ("searched_at", "TEXT"),
    ("created_at", "TEXT DEFAULT (datetime('now', 'localtime'))"),
]

EMPLOYEE_RANGES = [
    "1〜10名",
    "11〜50名",
    "51〜100名",
    "101〜300名",
    "301〜1000名",
    "1001〜5000名",
    "5001名以上",
]

OVERSEAS_OPTIONS = ["あり", "なし", "不明"]


def _get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """データベースとテーブルを初期化する（既存テーブルへのカラム追加も行う）"""
    conn = _get_connection()
    col_defs = ", ".join(f"{name} {typedef}" for name, typedef in COLUMNS)
    conn.execute(f"CREATE TABLE IF NOT EXISTS companies ({col_defs})")

    # 既存テーブルにカラムが無ければ追加（マイグレーション）
    existing = {row[1] for row in conn.execute("PRAGMA table_info(companies)").fetchall()}
    new_columns = ["product_genre", "employee_range", "has_overseas", "phone", "address"]
    for col in new_columns:
        if col not in existing:
            conn.execute(f"ALTER TABLE companies ADD COLUMN {col} TEXT DEFAULT ''")

    conn.commit()
    conn.close()


def add_company(data: dict) -> tuple[bool, str]:
    """企業を追加する。重複URLがある場合は警告を返す。

    Returns:
        (success: bool, message: str)
    """
    conn = _get_connection()

    # 重複チェック（URLベース）
    if data.get("url"):
        existing = conn.execute(
            "SELECT company_name FROM companies WHERE url = ?", (data["url"],)
        ).fetchone()
        if existing:
            conn.close()
            return False, f"同一URLの企業が既に登録されています: {existing['company_name']}"

    fields = ["company_name", "phone", "address", "url", "domain", "snippet",
              "industry", "product_genre", "prefecture", "employee_count",
              "employee_range", "revenue", "has_overseas", "notes", "searched_at"]
    values = [data.get(f, "") for f in fields]

    placeholders = ", ".join(["?"] * len(fields))
    field_names = ", ".join(fields)
    conn.execute(f"INSERT INTO companies ({field_names}) VALUES ({placeholders})", values)
    conn.commit()
    conn.close()
    return True, "登録完了"


def add_companies_bulk(companies: list[dict]) -> tuple[int, int, list[str]]:
    """企業を一括追加する。

    Returns:
        (added: int, skipped: int, messages: list[str])
    """
    added = 0
    skipped = 0
    messages = []

    for company in companies:
        success, msg = add_company(company)
        if success:
            added += 1
        else:
            skipped += 1
            messages.append(msg)

    return added, skipped, messages


def get_all_companies() -> pd.DataFrame:
    """全企業をDataFrameとして取得する"""
    conn = _get_connection()
    df = pd.read_sql_query("SELECT * FROM companies ORDER BY created_at DESC", conn)
    conn.close()
    return df


def get_companies_filtered(
    industry: str | None = None,
    prefecture: str | None = None,
    keyword: str | None = None,
    product_genre: str | None = None,
    employee_range: str | None = None,
    has_overseas: str | None = None,
) -> pd.DataFrame:
    """フィルター条件に基づいて企業を取得する"""
    conn = _get_connection()
    query = "SELECT * FROM companies WHERE 1=1"
    params = []

    if industry:
        query += " AND industry LIKE ?"
        params.append(f"%{industry}%")

    if prefecture:
        query += " AND prefecture LIKE ?"
        params.append(f"%{prefecture}%")

    if keyword:
        query += " AND (company_name LIKE ? OR snippet LIKE ? OR notes LIKE ?)"
        params.extend([f"%{keyword}%"] * 3)

    if product_genre:
        query += " AND product_genre LIKE ?"
        params.append(f"%{product_genre}%")

    if employee_range:
        query += " AND employee_range = ?"
        params.append(employee_range)

    if has_overseas:
        query += " AND has_overseas = ?"
        params.append(has_overseas)

    query += " ORDER BY created_at DESC"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def update_company(company_id: int, data: dict):
    """企業情報を更新する"""
    conn = _get_connection()
    fields = ["company_name", "phone", "address", "url", "domain", "snippet",
              "industry", "product_genre", "prefecture", "employee_count",
              "employee_range", "revenue", "has_overseas", "notes"]
    set_clause = ", ".join(f"{f} = ?" for f in fields if f in data)
    values = [data[f] for f in fields if f in data]
    values.append(company_id)

    conn.execute(f"UPDATE companies SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()


def delete_company(company_id: int):
    """企業を削除する"""
    conn = _get_connection()
    conn.execute("DELETE FROM companies WHERE id = ?", (company_id,))
    conn.commit()
    conn.close()


def delete_companies(company_ids: list[int]):
    """複数企業を一括削除する"""
    conn = _get_connection()
    placeholders = ", ".join(["?"] * len(company_ids))
    conn.execute(f"DELETE FROM companies WHERE id IN ({placeholders})", company_ids)
    conn.commit()
    conn.close()


def export_to_excel(df: pd.DataFrame) -> bytes:
    """DataFrameを装飾付きExcelファイルとしてバイト列で返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "企業リスト"

    # 表示用カラム（テレアポ用に電話番号・住所を会社名の次に配置）
    display_columns = {
        "company_name": "会社名",
        "phone": "電話番号",
        "address": "住所",
        "industry": "業種",
        "product_genre": "商品ジャンル",
        "prefecture": "所在地",
        "url": "URL",
        "snippet": "概要",
        "employee_count": "従業員数",
        "employee_range": "従業員数規模",
        "revenue": "売上規模",
        "has_overseas": "海外支店",
        "notes": "備考",
        "searched_at": "検索日時",
    }

    # 使用するカラムのみ
    cols = [c for c in display_columns if c in df.columns]
    headers = [display_columns[c] for c in cols]

    # ヘッダースタイル
    header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ヘッダー行
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # データ行
    data_font = Font(name="Yu Gothic", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(cols, 1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # 列幅の自動調整
    column_widths = {
        "company_name": 25,
        "phone": 18,
        "address": 35,
        "industry": 15,
        "product_genre": 18,
        "prefecture": 10,
        "url": 35,
        "snippet": 50,
        "employee_count": 12,
        "employee_range": 14,
        "revenue": 15,
        "has_overseas": 10,
        "notes": 30,
        "searched_at": 16,
    }
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(col_name, 15)

    # ヘッダー行を固定
    ws.freeze_panes = "A2"

    # オートフィルター
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

    # バイト列として出力
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# モジュールロード時にDB初期化
init_db()
