"""補助金レポート（Excel・HTMLメール本文）の生成"""

from __future__ import annotations

import io
from datetime import datetime
from html import escape

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_DISPLAY_COLUMNS = {
    "title": "補助金名",
    "source_category": "区分",
    "source_name": "発信元",
    "summary": "概要",
    "url": "URL",
    "matched_keyword": "ヒット条件",
    "domain": "ドメイン",
    "first_seen_at": "初回検出日時",
    "last_seen_at": "最終検出日時",
}


def build_excel_report(df: pd.DataFrame) -> bytes:
    """補助金一覧をExcelファイル（バイト列）として返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "週次補助金レポート"

    cols = [c for c in _DISPLAY_COLUMNS if c in df.columns]
    headers = [_DISPLAY_COLUMNS[c] for c in cols]

    header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    data_font = Font(name="Yu Gothic", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for r, (_, row) in enumerate(df.iterrows(), 2):
        for i, col in enumerate(cols, 1):
            val = row.get(col, "")
            c = ws.cell(row=r, column=i, value=val if pd.notna(val) else "")
            c.font = data_font
            c.alignment = data_align
            c.border = border
            if r % 2 == 0:
                c.fill = alt_fill

    widths = {
        "title": 40, "source_category": 10, "source_name": 22,
        "summary": 60, "url": 40, "matched_keyword": 20,
        "domain": 22, "first_seen_at": 18, "last_seen_at": 18,
    }
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 15)

    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_html_report(
    new_df: pd.DataFrame,
    all_df: pd.DataFrame,
    report_date: datetime | None = None,
) -> str:
    """メール本文用HTMLを生成する

    Args:
        new_df: 今週新たに検出された補助金
        all_df: DB上の累計補助金（参考統計用）
        report_date: レポート日時
    """
    report_date = report_date or datetime.now()

    by_category: dict[str, int] = {}
    if not new_df.empty and "source_category" in new_df.columns:
        for cat, group in new_df.groupby("source_category"):
            by_category[cat] = len(group)

    rows_html = []
    for _, row in new_df.iterrows():
        title = escape(str(row.get("title", "")))
        summary = escape(str(row.get("summary", "")))
        url = escape(str(row.get("url", "")))
        category = escape(str(row.get("source_category", "")))
        source = escape(str(row.get("source_name", "")))
        rows_html.append(
            f"""
            <tr>
                <td style="padding:8px;border:1px solid #ddd;vertical-align:top;">
                    <strong><a href="{url}" style="color:#2B5797;">{title}</a></strong><br>
                    <small style="color:#666;">{category} / {source}</small>
                </td>
                <td style="padding:8px;border:1px solid #ddd;vertical-align:top;font-size:13px;">
                    {summary}
                </td>
            </tr>
            """
        )

    category_summary = "".join(
        f"<li>{escape(cat)}: {n} 件</li>" for cat, n in sorted(by_category.items())
    ) or "<li>該当なし</li>"

    new_count = len(new_df)
    total_count = len(all_df)

    return f"""\
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>週次補助金レポート {report_date:%Y-%m-%d}</title>
</head>
<body style="font-family:'Yu Gothic',sans-serif;color:#222;line-height:1.6;">
<h2 style="color:#2B5797;border-bottom:2px solid #2B5797;padding-bottom:6px;">
週次 海外輸出関連 補助金レポート（{report_date:%Y-%m-%d}）
</h2>

<p>
本レポートは、日本国内の行政・金融機関・自治体・公的機関・営利団体が公開している
<strong>海外輸出に関連する補助金・助成金</strong>のうち、直近1週間で新たに検出されたものをまとめたものです。
</p>

<h3 style="color:#2B5797;">サマリー</h3>
<ul>
  <li>新規検出: <strong>{new_count}</strong> 件</li>
  <li>累計（DB上）: {total_count} 件</li>
</ul>
<p><strong>区分別の新規件数:</strong></p>
<ul>{category_summary}</ul>

<h3 style="color:#2B5797;">新規補助金一覧</h3>
{"<p>該当する新規補助金はありませんでした。</p>" if new_count == 0 else f'''
<table style="border-collapse:collapse;width:100%;font-size:14px;">
<thead>
<tr style="background:#2B5797;color:#fff;">
<th style="padding:8px;border:1px solid #2B5797;text-align:left;width:30%;">補助金名 / 発信元</th>
<th style="padding:8px;border:1px solid #2B5797;text-align:left;">概要</th>
</tr>
</thead>
<tbody>
{"".join(rows_html)}
</tbody>
</table>
'''}

<hr style="margin-top:24px;border:none;border-top:1px solid #ddd;">
<p style="color:#888;font-size:12px;">
このメールは GitHub Actions により毎週月曜 12:00（JST）に自動送信されています。<br>
収集ロジック・既定ソースは subsidy/sources.py で管理されています。
</p>
</body>
</html>
"""
